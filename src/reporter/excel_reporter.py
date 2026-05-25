"""
Excel 报表生成器
生成每日/每周舆情汇总报表，供内部审核使用
"""

import os
from datetime import datetime, date
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from loguru import logger


# 颜色定义
COLOR_HIGH_RISK = "FF4444"    # 红色 - 高风险
COLOR_MEDIUM_RISK = "FFA500"  # 橙色 - 中风险
COLOR_LOW_RISK = "44BB44"     # 绿色 - 低风险
COLOR_HEADER = "1E4080"       # 深蓝色 - 表头（中国移动蓝）
COLOR_HEADER_FONT = "FFFFFF"  # 白色字体


class ExcelReporter:
    """Excel 报表生成器"""

    def __init__(self, output_dir: str = "reports/output"):
        os.makedirs(output_dir, exist_ok=True)
        self.output_dir = output_dir

    def generate_daily_report(self, posts_data: List[dict], report_date: date = None) -> str:
        """
        生成每日舆情报表

        :param posts_data: 帖子数据列表（来自数据库查询）
        :param report_date: 报表日期，默认今天
        :return: 生成的文件路径
        """
        report_date = report_date or date.today()
        filename = f"反诈舆情日报_{report_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()

        # Sheet 1: 综合概览
        self._write_overview_sheet(wb.active, posts_data, report_date)
        wb.active.title = "综合概览"

        # Sheet 2: 高风险明细
        ws2 = wb.create_sheet("高风险明细")
        high_risk = [p for p in posts_data if p.get("risk_level") == "high"]
        self._write_detail_sheet(ws2, high_risk, "高风险舆情明细")

        # Sheet 3: 宁夏舆情明细（供客服/反诈专员研判解封）
        ws3 = wb.create_sheet("宁夏舆情明细")
        ningxia_posts = [p for p in posts_data if p.get("is_ningxia")]
        self._write_ningxia_sheet(ws3, ningxia_posts, "宁夏舆情研判清单")

        # Sheet 4: 全量数据
        ws4 = wb.create_sheet("全量数据")
        self._write_detail_sheet(ws4, posts_data, "全量舆情数据")

        # Sheet 5: 用户核查清单
        ws5 = wb.create_sheet("用户核查清单")
        self._write_user_check_sheet(ws5, posts_data)

        wb.save(filepath)
        logger.info(f"报表已生成: {filepath}")
        return filepath

    def _write_overview_sheet(self, ws, posts_data: List[dict], report_date: date):
        """写入概览 Sheet"""
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

        # 标题
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"客户自媒体反诈微舆情监测日报 — {report_date.strftime('%Y年%m月%d日')}"
        title_cell.font = Font(name="微软雅黑", size=16, bold=True, color=COLOR_HEADER_FONT)
        title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 统计摘要
        total = len(posts_data)
        high = sum(1 for p in posts_data if p.get("risk_level") == "high")
        medium = sum(1 for p in posts_data if p.get("risk_level") == "medium")
        low = sum(1 for p in posts_data if p.get("risk_level") == "low")
        suspected_fraud = sum(1 for p in posts_data if p.get("suspected_fraud_flag"))
        total_exposure = sum(p.get("view_count", 0) or 0 for p in posts_data)

        summary_data = [
            ("统计指标", "数值"),
            ("监测帖子总数", total),
            ("高风险舆情", high),
            ("中风险舆情", medium),
            ("低风险舆情", low),
            ("疑似诈骗用户施压", suspected_fraud),
            ("累计曝光量（估）", total_exposure),
            ("报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for row_idx, (label, value) in enumerate(summary_data, start=3):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        # 平台分布
        platform_col_start = 4
        ws.cell(row=3, column=platform_col_start, value="平台").font = Font(bold=True)
        ws.cell(row=3, column=platform_col_start + 1, value="条数").font = Font(bold=True)

        platform_counts = {}
        for p in posts_data:
            platform_counts[p.get("platform", "unknown")] = \
                platform_counts.get(p.get("platform", "unknown"), 0) + 1

        for i, (platform, count) in enumerate(sorted(platform_counts.items(), key=lambda x: -x[1])):
            ws.cell(row=4 + i, column=platform_col_start, value=platform)
            ws.cell(row=4 + i, column=platform_col_start + 1, value=count)

    def _write_detail_sheet(self, ws, posts_data: List[dict], title: str):
        """写入明细 Sheet"""
        columns = [
            ("序号", 6),
            ("平台", 12),
            ("风险等级", 10),
            ("标题/内容摘要", 45),
            ("作者", 15),
            ("粉丝数", 10),
            ("浏览量", 10),
            ("点赞数", 10),
            ("评论数", 10),
            ("转发数", 10),
            ("发布时间", 18),
            ("命中关键词", 25),
            ("疑似诈骗施压", 12),
            ("涉及手机号", 15),
            ("原文链接", 40),
        ]

        # 表头
        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.font = Font(color=COLOR_HEADER_FONT, bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 数据行
        risk_colors = {"high": COLOR_HIGH_RISK, "medium": COLOR_MEDIUM_RISK, "low": COLOR_LOW_RISK}
        risk_labels = {"high": "🔴 高风险", "medium": "🟡 中风险", "low": "🟢 低风险"}

        for row_idx, post in enumerate(posts_data, 2):
            risk = post.get("risk_level", "low")
            row_color = risk_colors.get(risk, COLOR_LOW_RISK)

            values = [
                row_idx - 1,
                post.get("platform", ""),
                risk_labels.get(risk, risk),
                post.get("title") or post.get("content", ""),  # 不截断，完整显示
                post.get("author_name", ""),
                post.get("author_followers", 0),
                post.get("view_count", 0),
                post.get("like_count", 0),
                post.get("comment_count", 0),
                post.get("share_count", 0),
                str(post.get("published_at", "")),  # 不截断，完整显示
                post.get("matched_keywords", ""),
                "是" if post.get("suspected_fraud_flag") else "否",
                post.get("mentioned_phones", ""),
                post.get("url", ""),
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 3:  # 风险等级列高亮
                    cell.fill = PatternFill("solid", fgColor=row_color)
                    cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _write_user_check_sheet(self, ws, posts_data: List[dict]):
        """写入用户核查清单 Sheet"""
        ws.cell(row=1, column=1, value="用户核查清单（含手机号）").font = Font(bold=True, size=14)

        columns = [
            ("手机号", 18), ("发现平台", 12), ("舆情风险", 10),
            ("发帖时间", 18), ("关停时间（待核查）", 20), ("核查状态", 12), ("备注", 30)
        ]
        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.font = Font(color=COLOR_HEADER_FONT, bold=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 提取含手机号的帖子
        row_idx = 3
        for post in posts_data:
            phones_raw = post.get("mentioned_phones", "")
            phones = phones_raw.split(",") if isinstance(phones_raw, str) and phones_raw else []
            if isinstance(phones_raw, list):
                phones = phones_raw

            for phone in phones:
                phone = phone.strip()
                if not phone:
                    continue
                ws.cell(row=row_idx, column=1, value=phone)
                ws.cell(row=row_idx, column=2, value=post.get("platform", ""))
                ws.cell(row=row_idx, column=3, value=post.get("risk_level", ""))
                ws.cell(row=row_idx, column=4, value=str(post.get("published_at", ""))[:16])
                ws.cell(row=row_idx, column=5, value="待核查")
                ws.cell(row=row_idx, column=6, value="待处理")
                ws.cell(row=row_idx, column=7, value="")
                row_idx += 1

    def _write_ningxia_sheet(self, ws, posts_data: List[dict], title: str):
        """写入宁夏舆情研判清单（供客服/反诈专员解封研判使用）"""
        # 标题行
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value = "宁夏反诈舆情研判清单 — 供客服/反诈专员解封研判使用"
        title_cell.font = Font(name="微软雅黑", size=14, bold=True, color=COLOR_HEADER_FONT)
        title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # 说明行
        ws.merge_cells("A2:L2")
        ws["A2"].value = "说明：本表汇总被识别为宁夏区域的反诈关停相关舆情，供研判用户是否可解封。重点关注：1) 是否有核验链接截图；2) 用户诉求合理性；3) 情感负面程度。"
        ws["A2"].font = Font(size=10, italic=True, color="666666")

        # 列定义（针对宁夏研判场景优化）
        columns = [
            ("序号", 6),
            ("平台", 10),
            ("风险等级", 10),
            ("用户手机号", 15),
            ("宁夏标识", 20),
            ("标题/内容摘要", 40),
            ("情感分值", 10),
            ("有核验链接", 10),
            ("链接类型", 15),
            ("发布时间", 18),
            ("原文链接", 35),
            ("研判建议", 30),
        ]

        # 表头（第3行）
        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.font = Font(color=COLOR_HEADER_FONT, bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 风险颜色和标签
        risk_colors = {"high": COLOR_HIGH_RISK, "medium": COLOR_MEDIUM_RISK, "low": COLOR_LOW_RISK}
        risk_labels = {"high": "🔴 高", "medium": "🟡 中", "low": "🟢 低"}

        # 数据行（从第4行开始）
        for row_idx, post in enumerate(posts_data, 4):
            risk = post.get("risk_level", "low")
            row_color = risk_colors.get(risk, COLOR_LOW_RISK)

            # 获取宁夏标识
            ningxia_ids = post.get("ningxia_identifiers", "")
            if isinstance(ningxia_ids, list):
                ningxia_ids = ", ".join(ningxia_ids)

            # 获取核验链接信息
            has_link = "是" if post.get("has_verification_link") else "否"
            link_types = post.get("verification_links_found", "")
            if isinstance(link_types, list):
                # 简化链接类型显示
                simplified = []
                for link in link_types:
                    if "online-cmcc" in link:
                        simplified.append("online-cmcc认证")
                    elif "nx.10086" in link:
                        simplified.append("宁夏10086认证")
                    else:
                        simplified.append(link[:20])
                link_types = ", ".join(simplified)

            # 情感分值（越低越负面）
            sentiment = post.get("sentiment_score", 0.5)
            sentiment_label = f"{sentiment:.2f}"
            if sentiment < 0.2:
                sentiment_label += " (极负面)"
            elif sentiment < 0.4:
                sentiment_label += " (负面)"

            # 生成研判建议
            suggestion = self._generate_ningxia_suggestion(post)

            # 获取手机号
            phones = post.get("ningxia_phones", "") or post.get("mentioned_phones", "")
            if isinstance(phones, list):
                phones = ", ".join(phones)

            values = [
                row_idx - 3,
                post.get("platform", ""),
                risk_labels.get(risk, risk),
                phones,
                ningxia_ids,
                post.get("title") or post.get("content", ""),
                sentiment_label,
                has_link,
                link_types,
                str(post.get("published_at", ""))[:16],
                post.get("url", ""),
                suggestion,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 3:  # 风险等级列高亮
                    cell.fill = PatternFill("solid", fgColor=row_color)
                    cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 添加底部说明
        last_row = len(posts_data) + 5
        ws.merge_cells(f"A{last_row}:L{last_row}")
        ws[f"A{last_row}"].value = "【研判要点】有核验链接截图的用户更可能是真实误关停，建议优先处理；情感极负面但无链接的需核实是否为施压用户。"
        ws[f"A{last_row}"].font = Font(size=10, bold=True, color="CC0000")

    def _generate_ningxia_suggestion(self, post: dict) -> str:
        """根据舆情特征生成研判建议"""
        suggestions = []

        # 有核验链接 → 可能是真实误关停用户
        if post.get("has_verification_link"):
            suggestions.append("有短信核验链接截图，可能是真实被关停用户，建议优先核查解封")

        # 情感极负面
        sentiment = post.get("sentiment_score", 0.5)
        if sentiment < 0.2:
            suggestions.append("情感极负面，用户情绪激动，需安抚处理")
        elif sentiment < 0.4:
            suggestions.append("情感偏负面，用户有不满情绪")

        # 有宁夏标识
        if post.get("ningxia_identifiers"):
            suggestions.append("明确提及宁夏移动/反诈中心，属本区责任范围")

        # 有手机号
        if post.get("ningxia_phones") or post.get("mentioned_phones"):
            suggestions.append("已提取手机号，可直接核查关停记录")

        # 疑似诈骗施压
        if post.get("suspected_fraud_flag"):
            suggestions.append("⚠️ 命中施压模式，需谨慎研判是否为真实用户")

        # 高传播
        if post.get("high_spread"):
            suggestions.append("传播量较大，需关注舆情发酵风险")

        return "；".join(suggestions) if suggestions else "常规舆情，按流程处理"
